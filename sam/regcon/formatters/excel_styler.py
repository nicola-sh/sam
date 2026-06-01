from __future__ import annotations

from pathlib import Path
from typing import Any

try:
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    import openpyxl
except ImportError:  # pragma: no cover

    def style_workbook(path: Path, excel_cfg: dict[str, Any]) -> None:
        return

else:

    def style_workbook(path: Path, excel_cfg: dict[str, Any]) -> None:
        wb = openpyxl.load_workbook(path)
        max_width = int(excel_cfg.get("max_column_width", 60))
        for sheet in wb.worksheets:
            for cell in sheet[1]:
                cell.font = Font(bold=True)
            for column_cells in sheet.columns:
                letter = get_column_letter(column_cells[0].column)
                length = max(
                    (len(str(c.value)) if c.value is not None else 0 for c in column_cells),
                    default=0,
                )
                sheet.column_dimensions[letter].width = min(length + 2, max_width)
            if excel_cfg.get("freeze_header", True):
                sheet.freeze_panes = "A2"
            if excel_cfg.get("auto_filter", True) and sheet.max_row > 1:
                sheet.auto_filter.ref = sheet.dimensions
        wb.save(path)
